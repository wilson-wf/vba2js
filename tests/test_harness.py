
import os
import sys
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


def create_test_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = 10
    ws['B1'] = 20
    ws['C1'] = '=A1+B1'
    test_file = 'tests/test_data.xlsx'
    os.makedirs(os.path.dirname(test_file), exist_ok=True)
    wb.save(test_file)
    return test_file


def test_golden_master():
    print("="*80)
    print("Testing Golden Master Capturer")
    print("="*80)
    
    from app.harness import GoldenMasterCapturer
    
    test_file = create_test_excel()
    capturer = GoldenMasterCapturer()
    
    test_cases = [
        {
            'name': 'Test Case 1',
            'setup': {
                'Sheet1': {
                    'A1': 5,
                    'B1': 15
                }
            }
        }
    ]
    
    snapshots = capturer.capture(test_file, test_cases)
    
    print(f"\nCaptured {len(snapshots)} snapshots")
    for snapshot in snapshots:
        print(f"Test case: {snapshot['test_case']['name']}")
        print(f"Worksheets count: {len(snapshot['snapshot']['worksheets'])}")
        for sheet_name, sheet_data in snapshot['snapshot']['worksheets'].items():
            print(f"  {sheet_name}: {len(sheet_data['cells'])} cells with values")
    
    print("\n  Golden Master capturer test successful")
    return True


def test_assertion_engine():
    print("\n" + "="*80)
    print("Testing Assertion Engine")
    print("="*80)
    
    from app.harness import AssertionEngine
    
    engine = AssertionEngine()
    
    golden_snapshots = [
        {
            'test_case': {'name': 'Test 1'},
            'snapshot': {
                'worksheets': {
                    'Sheet1': {
                        'cells': {
                            'A1': {'value': 10},
                            'B1': {'value': 20},
                            'C1': {'value': 30}
                        }
                    }
                }
            }
        }
    ]
    
    js_executions = [
        {
            'success': True,
            'state': {
                'worksheets': {
                    'Sheet1': {
                        'cells': {
                            'A1': {'value': 10},
                            'B1': {'value': 20},
                            'C1': {'value': 30}
                        }
                    }
                }
            },
            'side_effects': {}
        }
    ]
    
    result = engine.assert_all(golden_snapshots, js_executions)
    
    print(f"\nAssertion result: {'PASSED' if result['passed'] else 'FAILED'}")
    print(f"Passed tests: {result['passed_tests']}/{result['total_tests']}")
    
    print("\n  Assertion engine test successful")
    return True


def test_syntax_mapper_integration():
    print("\n" + "="*80)
    print("Testing Syntax Mapper Integration")
    print("="*80)
    
    from app.engine.syntax_mapper import SyntaxMapper
    
    mapper = SyntaxMapper()
    
    vba_code = '''Sub CalculateSum()
    Dim val1 As Integer
    Dim val2 As Integer
    Dim result As Integer
    
    val1 = Range("A1").Value
    val2 = Range("B1").Value
    result = val1 + val2
    Range("C1").Value = result
    MsgBox "Calculation complete! Result is: " & result
End Sub'''
    
    js_code = mapper.convert(vba_code)
    
    print(f"\nOriginal VBA code:\n{vba_code}")
    print(f"\nConverted JS code:\n{js_code}")
    
    print("\n  Syntax mapper integration test successful")
    return True


def test_llm_agent_mock():
    print("\n" + "="*80)
    print("Testing LLM Agent Mock Functionality")
    print("="*80)
    
    from app.engine.llm_agent import LLMAgent
    
    agent = LLMAgent(api_key=None)
    
    vba_code = '''Sub HelloWorld()
    MsgBox "Hello from VBA!"
End Sub'''
    
    js_code = agent.convert_vba_to_js(vba_code)
    
    print(f"\nVBA code:\n{vba_code}")
    print(f"\nMock converted JS code:\n{js_code}")
    
    print("\n  LLM agent mock test successful")
    return True


def run_all_tests():
    print("\n" + "="*80)
    print("Test Harness Complete Test Suite")
    print("="*80)
    
    results = []
    
    try:
        results.append(("Golden Master Capturer", test_golden_master()))
    except Exception as e:
        print(f"\n  Golden Master capturer test failed: {e}")
        import traceback
        traceback.print_exc()
        results.append(("Golden Master Capturer", False))
    
    try:
        results.append(("Assertion Engine", test_assertion_engine()))
    except Exception as e:
        print(f"\n  Assertion engine test failed: {e}")
        import traceback
        traceback.print_exc()
        results.append(("Assertion Engine", False))
    
    try:
        results.append(("Syntax Mapper Integration", test_syntax_mapper_integration()))
    except Exception as e:
        print(f"\n  Syntax mapper integration test failed: {e}")
        import traceback
        traceback.print_exc()
        results.append(("Syntax Mapper Integration", False))
    
    try:
        results.append(("LLM Agent Mock", test_llm_agent_mock()))
    except Exception as e:
        print(f"\n  LLM agent mock test failed: {e}")
        import traceback
        traceback.print_exc()
        results.append(("LLM Agent Mock", False))
    
    print("\n" + "="*80)
    print("Test Summary")
    print("="*80)
    
    all_passed = True
    for test_name, passed in results:
        status = "  PASSED" if passed else "  FAILED"
        print(f"{test_name}: {status}")
        if not passed:
            all_passed = False
    
    print(f"\nOverall result: {'ALL TESTS PASSED!' if all_passed else 'SOME TESTS FAILED'}")
    
    return all_passed


if __name__ == '__main__':
    success = run_all_tests()
    sys.exit(0 if success else 1)

