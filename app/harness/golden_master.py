
import shutil
import os
import uuid
from datetime import datetime
import openpyxl


class GoldenMasterCapturer:
    """Golden Master Capturer - VBA behavior snapshots
    
    Responsible for:
    1. Loading original Excel files
    2. Configuring test scenarios
    3. Capturing complete workbook states
    """
    
    def __init__(self, config=None):
        self.config = config
    
    def capture(self, xlsm_path, test_cases):
        """Complete capture of original VBA behavior"""
        snapshots = []
        
        for test_case in test_cases:
            print(f"  Capturing test scenario: {test_case.get('name', 'unnamed')}")
            
            # 1. Copy original file
            temp_file = self.create_temp_copy(xlsm_path)
            
            # 2. Initialize test scenario
            self.setup_test_scene(temp_file, test_case)
            
            # 3. Capture complete state (simulating VBA execution result)
            snapshot = self.capture_workbook_state(temp_file)
            
            # 4. Save snapshot
            snapshots.append({
                'test_case': test_case,
                'snapshot': snapshot
            })
            
            # Cleanup
            os.unlink(temp_file)
        
        return snapshots
    
    def create_temp_copy(self, source_path):
        """Create a temporary copy of the original file"""
        temp_dir = "data/temp"
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, f"{uuid.uuid4().hex}.xlsm")
        shutil.copy2(source_path, temp_path)
        return temp_path
    
    def setup_test_scene(self, file_path, test_case):
        """Initialize test scenario - set cell values"""
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
        
        setup_data = test_case.get('setup', {})
        for sheet_name, cell_values in setup_data.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            for cell_addr, value in cell_values.items():
                ws[cell_addr] = value
        
        wb.save(file_path)
        wb.close()
    
    def capture_workbook_state(self, file_path):
        """Capture complete workbook state"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        state = {
            'snapshot_id': str(uuid.uuid4()),
            'timestamp': datetime.now().isoformat(),
            'worksheets': {},
            'named_ranges': {},
            'defined_names': []
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            state['worksheets'][sheet_name] = self.capture_worksheet_state(ws)
        
        # Capture named ranges
        for name in wb.defined_names:
            state['named_ranges'][name.name] = name.attr_text
        
        wb.close()
        return state
    
    def capture_worksheet_state(self, worksheet):
        """Capture single worksheet state"""
        cells = {}
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_id = f"{cell.column_letter}{cell.row}"
                    cells[cell_id] = {
                        'value': cell.value,
                        'data_type': cell.data_type,
                        'number_format': cell.number_format
                    }
        
        return {
            'cells': cells,
            'row_dimensions': {
                str(r): worksheet.row_dimensions[r].height
                for r in worksheet.row_dimensions
            },
            'column_dimensions': {
                c: worksheet.column_dimensions[c].width
                for c in worksheet.column_dimensions
            }
        }

