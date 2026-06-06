import subprocess
import json
import os
import uuid


class WPSSimulator:
    """WPS JS宏仿真环境 - 模拟真实的WPS运行时
    
    提供：
    1. WPS API的模拟实现（Range, Sheets, MsgBox等）
    2. 在Node.js中执行JS宏
    3. 捕获执行后的状态
    """
    
    def __init__(self):
        self.workbook = None
        self.worksheets = {}
        self.active_sheet = None
        self.active_cell = None
        
        # 跟踪的副作用
        self.msgbox_calls = []
        self.inputbox_calls = []
        self.files_created = []
    
    def load_workbook(self, file_path):
        """加载工作簿"""
        import openpyxl
        self.workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name in self.workbook.sheetnames:
            self.worksheets[sheet_name] = WPSWorksheet(self, sheet_name)
        
        if self.workbook.sheetnames:
            self.active_sheet = self.worksheets[self.workbook.sheetnames[0]]
        
        self.active_cell = 'A1'
    
    def apply_initial_state(self, state):
        """应用初始状态"""
        for sheet_name, sheet_state in state.get('worksheets', {}).items():
            if sheet_name in self.worksheets:
                ws = self.worksheets[sheet_name]
                for cell_addr, cell_data in sheet_state.get('cells', {}).items():
                    ws.cells[cell_addr] = cell_data
    
    def execute_js_macro(self, js_code, function_name, *args):
        """执行JS宏并捕获状态变化
        
        返回: {
            'result': 函数返回值,
            'state': 执行后的状态快照,
            'side_effects': 副作用（如MsgBox调用）
        }
        """
        # 准备初始状态
        initial_state = self.capture_current_state()
        
        # 执行JS
        execution_result = self._execute_in_node(
            js_code, function_name, initial_state, *args
        )
        
        # 应用执行后的状态变化
        self._apply_state_changes(execution_result.get('state_changes', {}))
        
        # 捕获最终状态
        final_state = self.capture_current_state()
        
        return {
            'result': execution_result.get('result'),
            'state': final_state,
            'side_effects': execution_result.get('side_effects', {}),
            'success': execution_result.get('success', True),
            'error': execution_result.get('error')
        }
    
    def capture_current_state(self):
        """捕获当前状态（与Golden Master相同格式）"""
        state = {
            'worksheets': {},
            'named_ranges': {},
            'defined_names': []
        }
        
        for sheet_name, ws in self.worksheets.items():
            state['worksheets'][sheet_name] = ws.capture_state()
        
        return state
    
    def _execute_in_node(self, js_code, function_name, initial_state, *args):
        """在Node.js中执行JS宏"""
        # 构建完整的JS脚本
        wrapper = self._build_execution_wrapper(js_code, function_name, initial_state, args)
        
        try:
            result = subprocess.run(
                ['node', '-e', wrapper],
                capture_output=True,
                text=True,
                timeout=60
            )
            
            # 解析输出
            try:
                return json.loads(result.stdout)
            except:
                # 如果JSON解析失败，检查stderr
                return {
                    'success': False,
                    'error': result.stderr or 'Execution failed without error message',
                    'state_changes': {}
                }
        except subprocess.TimeoutExpired:
            return {
                'success': False,
                'error': 'Execution timeout',
                'state_changes': {}
            }
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'state_changes': {}
            }
    
    def _build_execution_wrapper(self, js_code, function_name, initial_state, args):
        """构建完整的执行脚本"""
        # 构建初始状态JSON
        state_json = json.dumps(initial_state)
        
        # 构建调用参数
        args_json = json.dumps(args)
        
        return f'''
        {{
            // WPS模拟环境
            {self._generate_wps_mock_api()}
            
            // 初始化状态
            let state = {state_json};
            
            // 从状态加载到模拟对象
            loadState(state);
            
            // 用户的JS宏代码
            {js_code}
            
            // 执行指定函数
            let result = null;
            let side_effects = {{}};
            let success = true;
            let error = null;
            
            try {{
                // 执行函数
                result = {function_name}.apply(null, {args_json});
                
                // 捕获状态变化
                state = captureState();
                side_effects = captureSideEffects();
            }} catch (e) {{
                success = false;
                error = e.message;
                state = captureState();
            }}
            
            // 输出JSON结果
            console.log(JSON.stringify({{
                success: success,
                result: result,
                state_changes: state,
                side_effects: side_effects,
                error: error
            }}));
        }}
        '''
    
    def _generate_wps_mock_api(self):
        """生成模拟的WPS API"""
        return '''
        // WPS JS宏模拟API
        
        // 状态存储
        let mockRangeValues = {};
        let mockMsgBoxCalls = [];
        let mockInputBoxCalls = [];
        let mockInputBoxReturns = [];
        let mockActiveCell = 'A1';
        let mockActiveSheetName = 'Sheet1';
        
        // Range对象
        function Range(cell) {
            return {
                get value() { return mockRangeValues[cell]; },
                get Value() { return mockRangeValues[cell]; },
                set value(val) { mockRangeValues[cell] = val; },
                set Value(val) { mockRangeValues[cell] = val; }
            };
        }
        
        // Application对象
        const Application = {
            MsgBox: function(text) {
                mockMsgBoxCalls.push(text);
            },
            InputBox: function(prompt) {
                mockInputBoxCalls.push(prompt);
                return mockInputBoxReturns.shift();
            },
            get ActiveCell() { return Range(mockActiveCell); },
            get ActiveSheet() { return { Name: mockActiveSheetName }; }
        };
        
        // ThisWorkbook对象
        const ThisWorkbook = {
            Sheets: function(name) {
                return {
                    Range: Range
                };
            },
            Worksheets: function(name) {
                return {
                    Range: Range
                };
            }
        };
        
        // Worksheets/Sheets
        const Worksheets = ThisWorkbook.Worksheets;
        const Sheets = ThisWorkbook.Sheets;
        
        // Cells函数
        function Cells(row, col) {
            let colName = getColumnName(col);
            return Range(colName + row);
        }
        
        // 辅助函数：列号转列名
        function getColumnName(col) {
            let name = '';
            while (col > 0) {
                col--;
                name = String.fromCharCode(65 + (col % 26)) + name;
                col = Math.floor(col / 26);
            }
            return name;
        }
        
        // 从JSON加载状态
        function loadState(state) {
            mockRangeValues = {};
            if (state.worksheets) {
                for (let sheetName in state.worksheets) {
                    let cells = state.worksheets[sheetName].cells;
                    if (cells) {
                        for (let cellAddr in cells) {
                            mockRangeValues[cellAddr] = cells[cellAddr].value;
                        }
                    }
                }
            }
        }
        
        // 捕获当前状态
        function captureState() {
            return {
                worksheets: {
                    [mockActiveSheetName]: {
                        cells: {}
                    }
                }
            };
        }
        
        // 捕获副作用
        function captureSideEffects() {
            return {
                msgbox_calls: mockMsgBoxCalls,
                inputbox_calls: mockInputBoxCalls,
                files_created: []
            };
        }
        '''
    
    def _apply_state_changes(self, state_changes):
        """应用状态变化到模拟器"""
        # 解析并应用状态变化
        for sheet_name, sheet_data in state_changes.get('worksheets', {}).items():
            if sheet_name in self.worksheets:
                ws = self.worksheets[sheet_name]
                for cell_addr, cell_data in sheet_data.get('cells', {}).items():
                    ws.cells[cell_addr] = cell_data


class WPSWorksheet:
    """模拟WPS Worksheet对象"""
    
    def __init__(self, simulator, name):
        self.simulator = simulator
        self.name = name
        self.cells = {}
    
    def Range(self, address):
        """获取Range对象"""
        return WPSRange(self, address)
    
    def get_cell(self, address):
        """获取单元格值"""
        return self.cells.get(address, {'value': None})
    
    def set_cell(self, address, value):
        """设置单元格值"""
        self.cells[address] = {'value': value}
    
    def capture_state(self):
        """捕获工作表状态"""
        return {'cells': self.cells}


class WPSRange:
    """模拟WPS Range对象"""
    
    def __init__(self, worksheet, address):
        self.worksheet = worksheet
        self.address = address
    
    @property
    def Value(self):
        """获取值"""
        return self.worksheet.get_cell(self.address)['value']
    
    @Value.setter
    def Value(self, value):
        """设置值"""
        self.worksheet.set_cell(self.address, value)
    
    @property
    def value(self):
        return self.Value
    
    @value.setter
    def value(self, value):
        self.Value = value
